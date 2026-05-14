import json
from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib.auth.decorators import login_required, user_passes_test
from django.urls import reverse_lazy
from .models import Student
from django.db.models import Q, Avg, Count

import openpyxl
from django.template.loader import get_template
from fpdf import FPDF

# --- RBAC UTILITY FUNCTIONS ---
def is_staff_user(user):
    """Check if user belongs to Teacher or Admin groups"""

    return user.groups.filter(name__in=['Teacher', 'Admin']).exists()

@login_required
@user_passes_test(is_staff_user) # Phase 1.1: Restrict dashboard to staff


def home(request):
    total_students = Student.objects.count()
    avg_cgpa_data = Student.objects.aggregate(Avg('cgpa'))
    avg_cgpa = round(avg_cgpa_data['cgpa__avg'] or 0, 2)
    dept_counts = Student.objects.values('department').annotate(total=Count('id'))

    context = {
        'total_students': total_students,
        'avg_cgpa': avg_cgpa,
        'dept_counts': dept_counts,
    }
    return render(request, 'students/home.html', context)
    

class StudentListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = Student
    template_name = 'students/student_list.html'
    context_object_name = 'students'
    paginate_by = 10

    def test_func(self): # Phase 1.1: Verification logic
        return is_staff_user(self.request.user)
    

    def get_queryset(self):
        
        queryset = super().get_queryset()
        
        
        query = self.request.GET.get('search')
        
        if query:
            # 3. Filter the list based on the search text
            queryset = queryset.filter(
                Q(user__icontains=query) | 
                Q(roll_number__icontains=query) |
                Q(department__icontains=query)
            )
        return queryset

    # This allows you to keep the search text in the input box after clicking search
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['query'] = self.request.GET.get('search', '')
        return context

class StudentDetailView(LoginRequiredMixin, UserPassesTestMixin, DetailView):
    model = Student
    template_name = 'students/student_detail.html'
    context_object_name = 'student'
    
    def test_func(self):
        return is_staff_user(self.request.user)

class StudentCreateView(LoginRequiredMixin,UserPassesTestMixin, CreateView):
    model = Student
    template_name = 'students/student_form.html'
    fields = ['user', 'roll_number', 'department', 'semester', 'cgpa', 'grade', 'phone', 'date_of_birth', 'address', 'profile_pic']
    success_url = reverse_lazy('student_list')

    def test_func(self): # Phase 1.1: Only Teachers/Admins can create
        return is_staff_user(self.request.user)



class StudentUpdateView(LoginRequiredMixin,UserPassesTestMixin, UpdateView):
    model = Student
    template_name = 'students/student_form.html'
    fields = ['user', 'roll_number', 'department', 'semester', 'cgpa', 'grade', 'phone', 'date_of_birth', 'address', 'profile_pic']
    success_url = reverse_lazy('student_list')

    def test_func(self): # Phase 1.1: Only Teachers/Admins can edit
        return is_staff_user(self.request.user)

class StudentDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    model = Student
    template_name = 'students/student_confirm_delete.html'
    success_url = reverse_lazy('student_list')

    def test_func(self): # Phase 1.1: High Priority - Restrict Deletion
        return self.request.user.groups.filter(name='Admin').exists() # Admin Only for Delete

@login_required
@user_passes_test(is_staff_user)
def export_students_excel(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="student_report.xlsx"'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Students"
    ws.append(['Roll Number', 'Name', 'Department', 'Semester', 'CGPA', 'Grade'])
    
    for s in Student.objects.all():
        ws.append([s.roll_number, s.user, s.department, s.semester, s.cgpa, s.grade])
    
    wb.save(response)
    return response
@login_required
@user_passes_test(is_staff_user)

def export_students_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="student_report.pdf"'
    
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", 'B', 16)
    pdf.cell(0, 10, 'Student Records Report', ln=True, align='C')
    pdf.ln(10)
    
    # Table Header
    pdf.set_font("Arial", 'B', 10)
    pdf.cell(35, 10, 'Roll No', 1)
    pdf.cell(50, 10, 'Name', 1)
    pdf.cell(40, 10, 'Dept', 1)
    pdf.cell(20, 10, 'CGPA', 1)
    pdf.cell(20, 10, 'Grade', 1)
    pdf.ln()
    
    # Table Body
    pdf.set_font("Arial", '', 10)
    for s in Student.objects.all():
        pdf.cell(35, 10, str(s.roll_number), 1)
        pdf.cell(50, 10, str(s.user)[:20], 1) # Trim name if too long
        pdf.cell(40, 10, str(s.department), 1)
        pdf.cell(20, 10, str(s.cgpa), 1)
        pdf.cell(20, 10, str(s.grade), 1)
        pdf.ln()
        
    response.write(pdf.output())
    return response


@login_required
def dashboard(request):
    # 1. Gather Global Stats
    total_students = Student.objects.count()
    avg_cgpa_data = Student.objects.aggregate(Avg('cgpa'))['cgpa__avg'] or 0
    
    # 2. Prepare Chart Data
    dept_data = Student.objects.values('department').annotate(count=Count('id'))
    labels = [item['department'] for item in dept_data]
    counts = [item['count'] for item in dept_data]

    perf_data = Student.objects.values('department').annotate(avg=Avg('cgpa'))
    perf_avgs = [round(item['avg'] or 0, 2) for item in perf_data]

    context = {
        'total_students': total_students,
        'avg_cgpa': round(avg_cgpa_data, 2),
        'dept_counts': dept_data,
        'labels': json.dumps(labels),
        'counts': json.dumps(counts),
        'perf_avgs': json.dumps(perf_avgs),
    }

    # 3. Role-Based Redirect
    if request.user.groups.filter(name='Student').exists():
        student_profile = Student.objects.filter(user_id=request.user.id).first()
        if student_profile:
            context['student'] = student_profile
            return render(request, 'students/student_dashboard.html', context)
    
    # Staff/Admin View
    return render(request, 'students/dashboard.html', context)